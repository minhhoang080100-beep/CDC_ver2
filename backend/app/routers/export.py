from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone
from app.core.database import db
from app.core.security import get_current_user
from app.core.permissions import can_manage_content
import io

router = APIRouter()


def _create_workbook():
    """Create openpyxl Workbook. Import here to avoid startup crash if not installed."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        return Workbook(), Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl chưa được cài đặt")


def _style_header(ws, headers, Font, PatternFill, Alignment):
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    return ws


@router.get("/surveys")
async def export_surveys(current_user: dict = Depends(get_current_user)):
    if not can_manage_content(current_user):
        raise HTTPException(status_code=403, detail="Bạn không có quyền xuất dữ liệu")

    wb, Font, PatternFill, Alignment = _create_workbook()
    ws = wb.active
    ws.title = "Khảo sát"
    headers = ["STT", "Tiêu đề", "Loại", "Trạng thái", "Số phản hồi", "Ngày tạo"]
    _style_header(ws, headers, Font, PatternFill, Alignment)

    surveys = await db.surveys.find({"isDeleted": {"$ne": True}}).sort("createdAt", -1).to_list(500)
    for i, s in enumerate(surveys, 1):
        response_count = await db.survey_responses.count_documents({"surveyId": str(s["_id"])})
        ws.append([
            i,
            s.get("title", ""),
            s.get("type", ""),
            s.get("status", ""),
            response_count,
            s.get("createdAt", "")
        ])

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=surveys_export.xlsx"}
    )


@router.get("/activities")
async def export_activities(current_user: dict = Depends(get_current_user)):
    if not can_manage_content(current_user):
        raise HTTPException(status_code=403, detail="Bạn không có quyền xuất dữ liệu")

    wb, Font, PatternFill, Alignment = _create_workbook()
    ws = wb.active
    ws.title = "Hoạt động"
    headers = ["STT", "Tiêu đề", "Thể loại", "Ngày diễn ra", "Đã đăng ký", "Đã tham gia", "Ngày tạo"]
    _style_header(ws, headers, Font, PatternFill, Alignment)

    activities = await db.activities.find({"isDeleted": {"$ne": True}}).sort("createdAt", -1).to_list(500)
    for i, a in enumerate(activities, 1):
        ws.append([
            i,
            a.get("name", ""),
            a.get("type", ""),
            str(a.get("time", "")),
            len(a.get("registrations", [])),
            len(a.get("attendances", [])),
            a.get("createdAt", "")
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=activities_export.xlsx"}
    )


@router.get("/users")
async def export_users(current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "SUPER_ADMIN":
        raise HTTPException(status_code=403, detail="Chỉ Super Admin mới có quyền xuất danh sách")

    wb, Font, PatternFill, Alignment = _create_workbook()
    ws = wb.active
    ws.title = "Người dùng"
    headers = ["STT", "Họ tên", "Tài khoản", "Mã Đoàn viên", "Phòng ban", "Vai trò", "Trạng thái"]
    _style_header(ws, headers, Font, PatternFill, Alignment)

    users = await db.users.find().sort("fullName", 1).to_list(1000)
    for i, u in enumerate(users, 1):
        ws.append([
            i,
            u.get("fullName", ""),
            u.get("username", ""),
            u.get("unionId", ""),
            u.get("department", ""),
            u.get("role", ""),
            u.get("status", "ACTIVE")
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users_export.xlsx"}
    )
